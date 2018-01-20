import openpyxl
import langdetect
from langdetect import detect

# set var to workbook
wb=openpyxl.load_workbook("vacatures.xlsx")

# Sheets in workbook
wb.get_sheet_names()

# set var to active sheet
ws=wb.active

#lang = detect(str_detect)

mylist = []

for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in col:
        mylist.append(cell.value)

        for item in mylist:
            print(item)
            language = detect(item)
            var = "De vacature in rij ", cell.row, "is in de taal: ", language
            print(var)

        #print(checklanguage)


    #lang = detect(cols)
    #print(lang)

# for i in range(1,ws.max_row):
#     if ws.cell(row=1, column=1).value == find_str:
#         for j in range(i, ws.max_column):
#             print (ws.cell(row=i, column=j).value)

