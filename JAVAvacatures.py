#Openpyxl commands
import langdetect
import bs4
import openpyxl
import datetime
import requests
import pprint
import re
import pyperclip
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup


# excel stuff
# set var to workbook
from langdetect import detect

urls = ['https://www.jouwictvacature.nl/vacatures/java', 'https://www.jouwictvacature.nl/vacatures/net','https://www.jouwictvacature.nl/vacatures/php', 'https://www.jouwictvacature.nl/vacatures/front-end-development']

#scrape sites
for url in urls:
    response = requests.get(url)
    page_soup = soup(response.content, "html.parser")

    containers = page_soup.findAll("article", {"class": "vacancy-item item up"})

    #loop through html
    for container in containers:

        # Datestamp
        date = datetime.date.today()

        # Developer
        if url == "https://www.jouwictvacature.nl/vacatures/java":
            developer = "Java Developer"
        if url == "https://www.jouwictvacature.nl/vacatures/net":
            developer = ".NET Developer"
        if url == "https://www.jouwictvacature.nl/vacatures/php":
            developer = "PHP Developer"
        if url == "https://www.jouwictvacature.nl/vacatures/front-end-development":
            developer = "Front-end Developer"


        #Company name
        company_container = container.div.h2
        company = company_container.text.strip()

        #Behandelaar

        leon = ["A", "B", "C", "D", "E", "F", "G", "H"]
        bram = ["I", "J", "K", "L", "M", "N", "O", "P"]
        renier = ["Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]

        for i in leon:
            if company.startswith(i):
                behandelaar = "Leon"

        for i in bram:
            if company.startswith(i):
                behandelaar = "Bram"

        for i in renier:
            if company.startswith(i):
                behandelaar = "Renier"

        #City
        city = container.div.a["title"]

        #URL
        full_url = container.div.div.a["href"]
        short_url = '=HYPERLINK("{}", "{}")'.format(full_url, "Link")

        #Title
        title_container = container.findAll("div", {"class":"info"})
        title = title_container[0].img["alt"]

        # Exp level
        find_words = ["Junior", "Medior", "Senior"]


        def find_str(s, char):
            index = 0

            if char in s:
                c = char[0]
                for ch in s:
                    if ch == c:
                        if s[index:index + len(char)] == char:
                            return index

                    index += 1

            return -1


        for i in find_words:
            result = find_str(title, i)
            if result == 0:
                explevel = i
                break
            else:
                explevel = ""


        #Description
        description = title_container[0].p.text
        language = detect(description)
        language_upper = language.upper()

        if language_upper == "NL":
            continue


        #Vervolgactie (Teamleader)
        teamleader = "Nieuw"

        # Connecties op LinkedIN?
        linkedin = "Nee"

        # Create list with var's
        companyinfo = [date,company,city,language_upper,explevel,developer,title,short_url,linkedin,teamleader,description,full_url]


        # Open Excel file
        wb = openpyxl.load_workbook('vacatures.xlsx')

        # Sheets in workbook
        wb.get_sheet_names()

        # set var to active sheet
        sheet = wb.active

        # define first empty row
        start_row = sheet.max_row + 1

        for y in container:

           for i in range(0,len(companyinfo)):
            e=sheet.cell(row=start_row,column=1+i)
            e.value=companyinfo[i]

        wb.save("vacatures.xlsx")
