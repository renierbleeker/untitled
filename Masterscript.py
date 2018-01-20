#Openpyxl commands
import langdetect
import bs4
import openpyxl
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup


# excel stuff
# set var to workbook
from langdetect import detect

wb=openpyxl.load_workbook('vacatures.xlsx')

# Sheets in workbook
wb.get_sheet_names()

# set var to active sheet
sheet=wb.active

# define first empty row and column
start_row = sheet.max_row + 1

#opening up connection grabbing the page
my_url = 'https://www.jouwictvacature.nl/vacatures/java'
uClient = uReq(my_url)
page_html = uClient.read()
uClient.close()

#HTML PARSING
page_soup = soup(page_html, "html.parser")

# grabs each vacancy
containers = page_soup.findAll("article", {"class":"vacancy-item item up"})

#write to excel
for container in containers:
    #Company name
    company_container = container.div.h2
    company = company_container.text.strip()

    #City
    city = container.div.a["title"]

    #URL
    url = container.div.div.a["href"]

    #Title
    title_container = container.findAll("div", {"class":"info"})
    title = title_container[0].img["alt"]

    #Description
    description = title_container[0].p.text
    language = detect(description)
    language_upper = language.upper()

    #Vervolgactie (Teamleader)
    vervolgactie = "Nieuw"

    companyinfo = [company, city,language_upper, title,description,url, vervolgactie]

    for y in container:

        for i in range(0,len(companyinfo)):
            e=sheet.cell(row=start_row,column=1+i)
            e.value=companyinfo[i]

wb.save("vacatures.xlsx")

