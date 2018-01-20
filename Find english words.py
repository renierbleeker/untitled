#Openpyxl commands
import bs4
import openpyxl
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup

# set var to workbook
wb=openpyxl.load_workbook("vacatures.xlsx")

# Sheets in workbook
wb.get_sheet_names()

# set var to active sheet
ws=wb.active

for i in range(1,ws.max_row):
    if ws.cell(row=1, column=1).value == "work":
        for j in range(i, ws.max_column):
            print (ws.cell(row=i, column=j).value)

