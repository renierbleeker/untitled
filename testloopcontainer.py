#Openpyxl commands
import bs4
import openpyxl
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup

# excel stuff
# set var to workbook
wb=openpyxl.load_workbook('vacatures.xlsx')

# Sheets in workbook
wb.get_sheet_names()

# set var to active sheet
sheet=wb.active

# define first empty row and column
start_row = sheet.max_row + 1

#opening up connection grabbing the page
my_url = 'https://www.jouwictvacature.nl/vacatures/java'
uClient = uReq(my_url)
page_html = uClient.read()
uClient.close()

#HTML PARSING
page_soup = soup(page_html, "html.parser")

# grabs each vacancy
containers = page_soup.findAll("article", {"class":"vacancy-item item up"})

containers_len = len(containers)
list_of_containers = list(range(10))
#print(containers)



# write to excel
# for items in range(0,containers_len):


for container in containers:
 # Company name
 company_container = container.div.h2
 company = company_container.text.strip()

 #City
 city = container.div.a["title"]

 #URL
 url = container.div.div.a["href"]

 #Title
 title_container = container.findAll("div", {"class":"info"})
 title = title_container[0].img["alt"]

 #Description
 description = title_container[0].p.text

 #Programming language
 hardskill = "Java"

 #Vervolgactie (Teamleader)
 vervolgactie = "Nieuw"

 companyinfo = [company, city, url, title, description, hardskill, vervolgactie]

 for y in container:

  for i in range(0,len(companyinfo)):
   e=sheet.cell(row=start_row,column=1+i)
   e.value=companyinfo[i]

wb.save("vacatures.xlsx")

