#Openpyxl commands
import langdetect
import bs4
import openpyxl
import datetime
import requests
import pprint
import re
import pyperclip
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup
from openpyxl import *

# Open Excel file
wb = openpyxl.load_workbook('vacatures ITouchables bram.xlsx')

# Sheets in workbook
wb.get_sheet_names()

# set var to active sheet
# ws = wb.active

# define first empty row
#start_row = sheet.max_row + 1

ws = wb.get_sheet_by_name('Vacatures')
find_word = "Uitgenodigd"

for i in range(1,ws.max_row):
    if ws.cell.value == find_word:
        for j in range(i, ws.max_column):
            print (ws.cell(row=i, column=j).value)